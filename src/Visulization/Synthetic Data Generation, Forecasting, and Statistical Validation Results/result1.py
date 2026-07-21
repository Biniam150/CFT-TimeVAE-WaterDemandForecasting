import pandas as pd
import numpy as np
import os
from pathlib import Path
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ==============================================================================
# CONFIGURATION
# ==============================================================================
root = Path(__file__).resolve().parents[3]
PROCESSED_DATA_DIR = root / "outputs" / "temporal_split_data"

CFT_VAE_RESULTS = root / "outputs" / "CFT-VAE" / "CFT-VAE-forecasting_results"
CFT_LSTM_TCN_RESULTS = CFT_VAE_RESULTS

TIME_VAE_RESULTS = root / "outputs" / "Time-VAE" / "Time-vae_forecasting_results"
TIME_LSTM_TCN_RESULTS = TIME_VAE_RESULTS

SAVE_VIS_DIR = root / "outputs" / "Visualization" / "Synthetic Data Generation, Forecasting, and Statistical Validation Results"
os.makedirs(SAVE_VIS_DIR, exist_ok=True)

LEVELS = [20, 70, 200, 700]
RATIO = 1.0

# ==============================================================================
# HELPERS
# ==============================================================================
def get_mean_demand(processed_dir, level, trial=1):
    train_path = os.path.join(processed_dir, "train", f"train_preprocessed_level_{level}_trial_{trial}.npz")
    if not os.path.exists(train_path):
        print(f"[!] Missing train file for level {level}")
        return None
    try:
        data = np.load(train_path, allow_pickle=True)
        feat_names = list(data["feature_names"])
        idx = feat_names.index("total_demand_clipped")
        norm_params = data["norm_params"].item()["total_demand_clipped"]
        val_norm = data["data"][:, :, idx]
        val_real = val_norm * (norm_params["max"] - norm_params["min"] + 1e-7) + norm_params["min"]
        return np.mean(val_real)
    except Exception as e:
        print(f"[!] Mean demand error for level {level}: {e}")
        return None

def format_mean_std(mean_val, std_val, decimals=3):
    if pd.isna(mean_val) or pd.isna(std_val):
        return ""
    return f"{mean_val:.{decimals}f}±{std_val:.{decimals}f}"

# ==============================================================================
# FLEXIBLE CONVERTER (This is the fix)
# ==============================================================================
def convert_lstm_tcn_to_long(deep_df):
    deep_df = deep_df.copy()
    deep_df.columns = [col.strip() for col in deep_df.columns]

    if {"training_type", "result_rmse"}.issubset(deep_df.columns):
        if "baseline_rmse" in deep_df.columns:
            return deep_df

        base_df = (
            deep_df[deep_df["training_type"] == "baseline"]
            [["level", "trial", "model", "result_rmse"]]
            .rename(columns={"result_rmse": "baseline_rmse"})
        )

        out = deep_df.merge(base_df, on=["level", "trial", "model"], how="left")
        return out[out["training_type"].isin(["synthetic_only", "augmented"])].copy()

    syn_col = None
    for c in ["synthetic_alone", "synthetic_only", "synth_alone", "syn", "synthetic"]:
        if c in deep_df.columns:
            syn_col = c
            break

    rows = []
    for _, row in deep_df.iterrows():
        base = row.get("base", np.nan)

        # Synthetic Only
        if syn_col and pd.notna(row.get(syn_col)):
            rows.append({
                "level": row["level"], "trial": row["trial"], "model": row["model"],
                "training_type": "synthetic_only",
                "baseline_rmse": base,
                "result_rmse": row[syn_col]
            })

        # Augmented
        if pd.notna(row.get("aug")):
            rows.append({
                "level": row["level"], "trial": row["trial"], "model": row["model"],
                "training_type": "augmented",
                "baseline_rmse": base,
                "result_rmse": row["aug"]
            })
    return pd.DataFrame(rows)

# ==============================================================================
# LOAD CFT-VAE & TIMEVAE
# ==============================================================================
def load_cft_level_results(level):
    ratio_str = str(RATIO).replace(".", "p")
    dfs = []
    cl_path = os.path.join(CFT_VAE_RESULTS, f"results_level{level}_ratio{ratio_str}.csv")
    deep_path = os.path.join(CFT_LSTM_TCN_RESULTS, f"results_level{level}_ratio{ratio_str}_lstm_tcn.csv")
    if os.path.exists(cl_path): dfs.append(pd.read_csv(cl_path))
    if os.path.exists(deep_path): dfs.append(convert_lstm_tcn_to_long(pd.read_csv(deep_path)))
    return pd.concat(dfs, ignore_index=True) if dfs else None

def load_timevae_level_results(level):
    ratio_str = str(RATIO).replace(".", "p")
    dfs = []
    cl_path = os.path.join(TIME_VAE_RESULTS, f"results_level{level}_ratio{ratio_str}.csv")
    deep_path = os.path.join(TIME_LSTM_TCN_RESULTS, f"results_lvl{level}_lstm_tcn_original_timevae_ratio{ratio_str}.csv")
    if os.path.exists(cl_path): dfs.append(pd.read_csv(cl_path))
    if os.path.exists(deep_path): dfs.append(convert_lstm_tcn_to_long(pd.read_csv(deep_path)))
    return pd.concat(dfs, ignore_index=True) if dfs else None

# ==============================================================================
# METRICS
# ==============================================================================
def get_metrics_raw(df, model_name, mean_dem):
    if df is None or df.empty:
        return {"base_m":np.nan,"base_s":np.nan,"syn_m":np.nan,"syn_s":np.nan,"aug_m":np.nan,"aug_s":np.nan}

    m_df = df[df["model"] == model_name].copy()
    if m_df.empty:
        return {"base_m":np.nan,"base_s":np.nan,"syn_m":np.nan,"syn_s":np.nan,"aug_m":np.nan,"aug_s":np.nan}

    base_nrmse = m_df.drop_duplicates(subset=["trial"])["baseline_rmse"].values / mean_dem
    syn_nrmse = m_df[m_df["training_type"]=="synthetic_only"]["result_rmse"].values / mean_dem
    aug_nrmse = m_df[m_df["training_type"]=="augmented"]["result_rmse"].values / mean_dem

    return {
        "base_m": np.mean(base_nrmse), "base_s": np.std(base_nrmse),
        "syn_m": np.mean(syn_nrmse) if len(syn_nrmse)>0 else np.nan,
        "syn_s": np.std(syn_nrmse) if len(syn_nrmse)>0 else np.nan,
        "aug_m": np.mean(aug_nrmse) if len(aug_nrmse)>0 else np.nan,
        "aug_s": np.std(aug_nrmse) if len(aug_nrmse)>0 else np.nan,
    }

# ==============================================================================
# MAIN
# ==============================================================================
def run_compiled_analysis():
    print("\n" + "="*120)
    print("WR2 COMPARATIVE ANALYSIS: TimeVAE vs CFT-VAE".center(120))
    print("="*120)

    model_order = ["GradientBoosting", "RandomForest", "SVM", "XGBoost", "LSTM", "TCN"]
    level_tables = []

    for level in LEVELS:
        mean_dem = get_mean_demand(PROCESSED_DATA_DIR, level)
        if mean_dem is None: continue

        df_cft = load_cft_level_results(level)
        df_time = load_timevae_level_results(level)

        models = [m for m in model_order if any(
            d is not None and m in d["model"].values
            for d in [df_cft, df_time]
        )]

        print(f"\nLEVEL {level} | Mean Demand: {mean_dem:.2f} m³/h")
        print(f"{'Model':<18} | {'Baseline':<15} | TimeVAE Syn | TimeVAE Aug | CFT Syn | CFT Aug")
        print("-"*120)

        level_rows = []

        for model in models:
            c = get_metrics_raw(df_cft, model, mean_dem)
            t = get_metrics_raw(df_time, model, mean_dem)

            baseline = format_mean_std(c["base_m"], c["base_s"]) or format_mean_std(t["base_m"], t["base_s"])
            time_syn = format_mean_std(t["syn_m"], t["syn_s"])
            time_aug = format_mean_std(t["aug_m"], t["aug_s"])
            cft_syn = format_mean_std(c["syn_m"], c["syn_s"])
            cft_aug = format_mean_std(c["aug_m"], c["aug_s"])

            print(f"{model:<18} | {baseline:<15} | "
                  f"{time_syn:<12} | {time_aug:<12} | "
                  f"{cft_syn:<12} | {cft_aug:<12}")

            level_rows.append({
                "Model": model,
                "Baseline": baseline,
                "TimeVAE Syn": time_syn,
                "TimeVAE Aug": time_aug,
                "CFT Syn": cft_syn,
                "CFT Aug": cft_aug,
            })

        level_tables.append((level, mean_dem, pd.DataFrame(level_rows)))

    if level_tables:
        output_path = os.path.join(SAVE_VIS_DIR, "Comparison_TimeVAE_CFT.xlsx")
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            start_row = 0
            for level, mean_dem, table in level_tables:
                table.to_excel(
                    writer,
                    sheet_name="Comparison",
                    index=False,
                    startrow=start_row + 2,
                )
                worksheet = writer.sheets["Comparison"]
                title_row = start_row + 1
                worksheet.merge_cells(
                    start_row=title_row,
                    start_column=1,
                    end_row=title_row,
                    end_column=6,
                )
                title_cell = worksheet.cell(title_row, 1)
                title_cell.value = f"LEVEL {level} | Mean Demand: {mean_dem:.2f} m³/h"
                title_cell.fill = PatternFill("solid", fgColor="1F4E78")
                title_cell.font = Font(color="FFFFFF", bold=True, size=12)
                title_cell.alignment = Alignment(horizontal="left", vertical="center")
                worksheet.row_dimensions[title_row].height = 24

                header_row = start_row + 3
                for cell in worksheet[header_row]:
                    cell.fill = PatternFill("solid", fgColor="D9EAF7")
                    cell.font = Font(bold=True, color="1F1F1F")
                    cell.alignment = Alignment(horizontal="center")
                    cell.border = Border(bottom=Side(style="thin", color="7F8C8D"))

                first_data_row = header_row + 1
                last_data_row = header_row + len(table)
                for row in worksheet.iter_rows(
                    min_row=first_data_row,
                    max_row=last_data_row,
                    min_col=2,
                    max_col=6,
                ):
                    for cell in row:
                        cell.alignment = Alignment(horizontal="center")

                start_row = last_data_row + 2

            worksheet = writer.sheets["Comparison"]
            worksheet.sheet_view.showGridLines = False
            worksheet.freeze_panes = "A4"
            widths = {"A": 24, "B": 18, "C": 18, "D": 18, "E": 16, "F": 16}
            for column, width in widths.items():
                worksheet.column_dimensions[column].width = width

        print(f"\nSaved successfully: {output_path}")

if __name__ == "__main__":
    run_compiled_analysis()
