import gc
import os
import re
import time
from pathlib import Path

os.environ["TF_CPP_MIN_LOG_LEVEL"] = "2"
os.environ["TF_DETERMINISTIC_OPS"] = "1"
os.environ.setdefault("PROTOCOL_BUFFERS_PYTHON_IMPLEMENTATION", "python")

import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sklearn.ensemble import GradientBoostingRegressor, RandomForestRegressor
from sklearn.metrics import mean_squared_error
from sklearn.preprocessing import StandardScaler
from sklearn.svm import SVR
from xgboost import XGBRegressor

try:
    import psutil
except ImportError:
    psutil = None


LEVEL = 700
RATIO = 1.0
RATIO_SUFFIX = "1p0"
RANDOM_STATE = 42

REPO_ROOT = Path(__file__).resolve().parents[4]
PROCESSED_DIR = REPO_ROOT / "outputs" / "temporal_split_data"
SYNTHETIC_DIR = (
    REPO_ROOT / "outputs" / "CFT-VAE" / "synthetic_data" / f"ratio_{RATIO_SUFFIX}"
)
SAVE_DIR = (
    REPO_ROOT
    / "outputs"
    / "Visualization"
    / "Computational Performance and Scalability Analysis"
    / "Forecasting"
)

MODEL_ORDER = [
    "RandomForest",
    "XGBoost",
    "GradientBoosting",
    "SVM",
    "LSTM",
    "TCN",
]


def get_process_memory_mb():
    if psutil is None:
        return np.nan
    return psutil.Process(os.getpid()).memory_info().rss / (1024 ** 2)


def load_real_data(path):
    with np.load(path, allow_pickle=True) as data:
        return {
            "data": data["data"].astype(np.float32),
            "norm_params": data["norm_params"].item(),
        }


def load_synthetic_data(path):
    with np.load(path, allow_pickle=True) as data:
        return data["synthetic_samples"].astype(np.float32)


def denormalize_data(data, norm_params, feature_name="total_demand_clipped"):
    norm_info = norm_params[feature_name]
    return data * (norm_info["max"] - norm_info["min"] + 1e-7) + norm_info["min"]


def create_features_and_targets(time_series_data):
    if time_series_data.ndim == 3:
        flattened = time_series_data[:, :, 0].flatten()
    else:
        flattened = time_series_data.flatten()

    time_index = pd.date_range(start="2020-01-01", periods=len(flattened), freq="h")
    demand = pd.DataFrame({"demand": flattened}, index=time_index)
    features = pd.DataFrame(index=time_index)

    for lag in [1, 2, 3, 24, 25, 48, 168]:
        features[f"lag_{lag}"] = demand["demand"].shift(lag)

    features["roll_24h_mean"] = demand["demand"].shift(1).rolling(24, min_periods=12).mean()
    features["roll_24h_std"] = demand["demand"].shift(1).rolling(24, min_periods=12).std()
    features["roll_168h_mean"] = demand["demand"].shift(1).rolling(168, min_periods=24).mean()
    features["hour"] = features.index.hour
    features["dayofweek"] = features.index.dayofweek
    features["month"] = features.index.month
    features["hour_sin"] = np.sin(2 * np.pi * features.index.hour / 24)
    features["hour_cos"] = np.cos(2 * np.pi * features.index.hour / 24)
    features["dayofweek_sin"] = np.sin(2 * np.pi * features.index.dayofweek / 7)
    features["dayofweek_cos"] = np.cos(2 * np.pi * features.index.dayofweek / 7)

    combined = pd.concat([features, demand["demand"].rename("target")], axis=1).dropna()
    return combined.drop(columns="target"), combined["target"]


class NeuralRegressor:
    def __init__(self, architecture, epochs=30, batch_size=32):
        self.architecture = architecture
        self.epochs = epochs
        self.batch_size = batch_size
        self.scaler = StandardScaler()
        self.model = None
        self.tf = None

    def fit(self, X, y):
        import tensorflow as tf
        from tensorflow.keras.callbacks import EarlyStopping
        from tensorflow.keras.layers import Conv1D, Dense, Flatten, LSTM
        from tensorflow.keras.models import Sequential

        self.tf = tf
        X_scaled = self.scaler.fit_transform(X)
        X_arr = X_scaled.reshape(X_scaled.shape[0], X_scaled.shape[1], 1).astype(np.float32)
        y_arr = y.to_numpy(dtype=np.float32)

        if self.architecture == "LSTM":
            self.model = Sequential([
                LSTM(64, input_shape=(X.shape[1], 1)),
                Dense(32, activation="relu"),
                Dense(1),
            ])
            callbacks = [
                EarlyStopping(
                    monitor="val_loss",
                    patience=10,
                    restore_best_weights=True,
                    verbose=0,
                )
            ]
            fit_kwargs = {"validation_split": 0.1, "shuffle": True, "callbacks": callbacks}
        else:
            self.model = Sequential([
                Conv1D(
                    filters=64,
                    kernel_size=3,
                    dilation_rate=1,
                    padding="causal",
                    activation="relu",
                    input_shape=(X.shape[1], 1),
                ),
                Conv1D(
                    filters=64,
                    kernel_size=3,
                    dilation_rate=2,
                    padding="causal",
                    activation="relu",
                ),
                Flatten(),
                Dense(32, activation="relu"),
                Dense(1),
            ])
            fit_kwargs = {}

        self.model.compile(optimizer="adam", loss="mse")
        self.model.fit(
            X_arr,
            y_arr,
            epochs=self.epochs,
            batch_size=self.batch_size,
            verbose=0,
            **fit_kwargs,
        )

    def predict(self, X):
        X_scaled = self.scaler.transform(X)
        X_arr = X_scaled.reshape(X_scaled.shape[0], X_scaled.shape[1], 1).astype(np.float32)
        return self.model.predict(X_arr, verbose=0).flatten()

    def clear(self):
        if self.tf is not None:
            self.tf.keras.backend.clear_session()


def create_model(name):
    if name == "RandomForest":
        return RandomForestRegressor(
            n_estimators=100,
            max_depth=10,
            random_state=RANDOM_STATE,
            n_jobs=-1,
        )
    if name == "XGBoost":
        return XGBRegressor(
            n_estimators=100,
            learning_rate=0.05,
            max_depth=6,
            random_state=RANDOM_STATE,
            n_jobs=-1,
        )
    if name == "GradientBoosting":
        return GradientBoostingRegressor(
            n_estimators=100,
            learning_rate=0.05,
            max_depth=5,
            random_state=RANDOM_STATE,
        )
    if name == "SVM":
        return SVR(kernel="rbf", C=100, gamma="scale", epsilon=0.1)
    if name in {"LSTM", "TCN"}:
        return NeuralRegressor(name, epochs=30, batch_size=32)
    raise ValueError(f"Unknown model: {name}")


def discover_available_trials():
    trials = []
    train_dir = PROCESSED_DIR / "train"
    for path in train_dir.glob(f"train_preprocessed_level_{LEVEL}_trial_*.npz"):
        match = re.search(r"trial_(\d+)", path.stem)
        if not match:
            continue
        trial = int(match.group(1))
        test_path = PROCESSED_DIR / "test" / f"test_preprocessed_level_{LEVEL}_trial_{trial}.npz"
        syn_path = SYNTHETIC_DIR / f"synthetic_level{LEVEL}_trial{trial}_ratio{RATIO_SUFFIX}.npz"
        if test_path.exists() and syn_path.exists():
            trials.append(trial)
    return sorted(trials)


def profile_model(model, X_train, y_train, X_test):
    start_train = time.perf_counter()
    model.fit(X_train, y_train)
    training_time = time.perf_counter() - start_train
    memory_mb = get_process_memory_mb()

    start_inference = time.perf_counter()
    predictions = model.predict(X_test)
    inference_latency = time.perf_counter() - start_inference
    return predictions, training_time, memory_mb, inference_latency


def save_excel(results, trial, output_path):
    table = pd.DataFrame(results)[
        ["Model", "Dataset", "Training time [s]", "Memory [MB]", "Inference latency [s]"]
    ]

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        table.to_excel(writer, sheet_name="Computational Profile", index=False, startrow=2)
        worksheet = writer.sheets["Computational Profile"]

        worksheet.merge_cells("A1:E1")
        title = worksheet["A1"]
        title.value = f"Level {LEVEL} | Trial {trial} | Full Conditioning | Synthetic Ratio 1:1"
        title.fill = PatternFill("solid", fgColor="1F4E78")
        title.font = Font(color="FFFFFF", bold=True, size=12)
        title.alignment = Alignment(horizontal="center", vertical="center")
        worksheet.row_dimensions[1].height = 25

        for cell in worksheet[3]:
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(bottom=Side(style="thin", color="7F8C8D"))
        worksheet.row_dimensions[3].height = 34

        start_row = 4
        for model_index, _ in enumerate(MODEL_ORDER):
            first_row = start_row + model_index * 2
            second_row = first_row + 1
            worksheet.merge_cells(start_row=first_row, start_column=1, end_row=second_row, end_column=1)
            worksheet.cell(first_row, 1).alignment = Alignment(horizontal="center", vertical="center")
            for row in [first_row, second_row]:
                for column in range(2, 6):
                    worksheet.cell(row, column).alignment = Alignment(horizontal="center")
            for column in range(1, 6):
                worksheet.cell(second_row, column).border = Border(
                    bottom=Side(style="thin", color="7F8C8D")
                )

        for row in range(4, 4 + len(table)):
            worksheet.cell(row, 3).number_format = "0.00"
            worksheet.cell(row, 4).number_format = "0.00"
            worksheet.cell(row, 5).number_format = "0.0000"

        widths = {"A": 22, "B": 16, "C": 20, "D": 16, "E": 22}
        for column, width in widths.items():
            worksheet.column_dimensions[column].width = width
        worksheet.sheet_view.showGridLines = False
        worksheet.freeze_panes = "A4"


def run_profile():
    total_start = time.perf_counter()
    np.random.seed(RANDOM_STATE)

    available_trials = discover_available_trials()
    if not available_trials:
        raise FileNotFoundError("No complete Level 700 real/synthetic trial files were found.")

    print(f"Available Level {LEVEL} trials: {available_trials}")
    selected_trial = int(input("Enter the trial number to profile: "))
    if selected_trial not in available_trials:
        raise ValueError(f"Trial {selected_trial} is not available for Level {LEVEL}.")

    train_path = PROCESSED_DIR / "train" / f"train_preprocessed_level_{LEVEL}_trial_{selected_trial}.npz"
    test_path = PROCESSED_DIR / "test" / f"test_preprocessed_level_{LEVEL}_trial_{selected_trial}.npz"
    synthetic_path = (
        SYNTHETIC_DIR
        / f"synthetic_level{LEVEL}_trial{selected_trial}_ratio{RATIO_SUFFIX}.npz"
    )

    train_data = load_real_data(train_path)
    test_data = load_real_data(test_path)
    synthetic = load_synthetic_data(synthetic_path)

    train_real = denormalize_data(train_data["data"][:, :, 0], train_data["norm_params"])
    test_real = denormalize_data(test_data["data"][:, :, 0], test_data["norm_params"])
    X_train, y_train = create_features_and_targets(train_real)
    X_test, y_test = create_features_and_targets(test_real)
    X_syn, y_syn = create_features_and_targets(synthetic)

    common_features = X_train.columns.intersection(X_syn.columns)
    X_train = X_train[common_features]
    X_test = X_test[common_features]
    X_syn = X_syn[common_features]
    X_augmented = pd.concat([X_train, X_syn], ignore_index=True)
    y_augmented = pd.concat([y_train, y_syn], ignore_index=True)

    print(f"\nUsing {len(common_features)} engineered features.")
    print(f"Synthetic file: {synthetic_path}")

    scenarios = [
        ("Baseline", X_train, y_train),
        ("Augmented", X_augmented, y_augmented),
    ]
    results = []

    for model_name in MODEL_ORDER:
        print(f"\n{model_name}")
        for dataset_name, X_scenario, y_scenario in scenarios:
            if model_name in {"LSTM", "TCN"}:
                import tensorflow as tf

                tf.random.set_seed(RANDOM_STATE)

            model = create_model(model_name)
            predictions, training_time, memory_mb, inference_latency = profile_model(
                model, X_scenario, y_scenario, X_test
            )
            rmse = float(np.sqrt(mean_squared_error(y_test, predictions)))

            results.append({
                "Model": model_name,
                "Dataset": dataset_name,
                "Training time [s]": training_time,
                "Memory [MB]": memory_mb,
                "Inference latency [s]": inference_latency,
                "RMSE": rmse,
            })
            print(
                f"  {dataset_name:<10} | Train: {training_time:.2f} s | "
                f"Memory: {memory_mb:.2f} MB | Inference: {inference_latency:.4f} s"
            )

            if isinstance(model, NeuralRegressor):
                model.clear()
            del model, predictions
            gc.collect()

    output_table = pd.DataFrame(results)[
        ["Model", "Dataset", "Training time [s]", "Memory [MB]", "Inference latency [s]"]
    ]
    print("\n" + "=" * 90)
    print(output_table.to_string(index=False))
    print("=" * 90)

    SAVE_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = time.strftime("%Y%m%d_%H%M%S")
    output_path = SAVE_DIR / (
        f"computational_profile_6models_level{LEVEL}_trial{selected_trial}_{timestamp}.xlsx"
    )
    save_excel(results, selected_trial, output_path)

    elapsed = time.perf_counter() - total_start
    print(f"\nExcel file saved to:\n{output_path}")
    print(f"Total profiling time: {elapsed / 60:.2f} minutes")


if __name__ == "__main__":
    run_profile()
